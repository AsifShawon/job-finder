import csv
import hashlib
import io
import re
import unicodedata
from datetime import UTC, datetime, timedelta
from urllib.parse import urlparse

import logging
from fastapi import APIRouter, Depends, File, Form, HTTPException, Response, UploadFile
from pydantic import BaseModel, Field
from sqlalchemy import case, delete, func, select, text
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from app.core.deps import get_admin_user, get_db
from app.ingestion.eligibility_engine import tag_eligibility
from app.ingestion.extractor import extract_structured
from app.ingestion.official_sources import ensure_official_sources
from app.ingestion.scrapeability import check_scrapeability
from app.ingestion.validators import parse_deadline
from app.models.entities import (
    AlertEvent,
    AlertRule,
    CrawlJob,
    CrawlRun,
    Opportunity,
    RawDocument,
    Source,
    User,
)
from app.models.enums import CrawlRunStatus, CrawlStatus
from app.schemas.admin import (
    AdminAiSettingsOut,
    AdminAiSettingsUpdate,
    AdminOverviewOut,
    AdminOverviewStats,
    AdminSourceOut,
    BulkImportError,
    BulkImportResult,
    CrawlJobOut,
    CrawlJobPage,
    CrawlRunOut,
    CrawlRunPage,
    DataResetResult,
    FailedExtractionOut,
    FailedExtractionPage,
    ManualEntryPrefillRequest,
    ManualEntryPrefillResponse,
    ManualJobEntryRequest,
    ManualEntryBulkImportResult,
    ReviewQueueOut,
    ReviewQueuePage,
    ReviewStatusUpdate,
    SourceProbeRequest,
    SourceProbeResult,
    SourceTestResult,
    TriggerAllResult,
)
from app.schemas.common import MessageResponse
from app.schemas.opportunity import RawDocumentOut, RawDocumentPage
from app.schemas.source import SourceCreate, SourceOut, SourceUpdate
from app.services.runtime_settings_service import (
    AI_API_KEY,
    AI_MODEL,
    AI_PROVIDER,
    get_ai_api_key,
    get_ai_model,
    get_ai_provider,
    has_ai_api_key,
    set_setting,
)

router = APIRouter(prefix="/admin", tags=["admin"])
logger = logging.getLogger(__name__)
STALE_RUNNING_CRAWL_MINUTES = 30
MANUAL_SOURCE_NAME = "Manual Entry"
MANUAL_SOURCE_URL = "https://manual-entry.local/"


# ── Helpers ────────────────────────────────────────────────────────────────────

def _slugify(text: str) -> str:
    text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")
    text = re.sub(r"[^\w\s-]", "", text.lower())
    return re.sub(r"[-\s]+", "-", text).strip("-")[:580]


def _refresh_draft_search_tsv(db: Session, draft_id: int) -> None:
    db.execute(
        text(
            "UPDATE opportunities "
            "SET search_tsv = to_tsvector('simple', "
            "coalesce(title,'') || ' ' || coalesce(summary_bn,'') || ' ' || "
            "coalesce(summary_en,'') || ' ' || coalesce(eligibility_text,'')) "
            "WHERE id = :id"
        ),
        {"id": draft_id},
    )


def _manual_source_item_key(source_url: str) -> str:
    return hashlib.sha256(source_url.encode()).hexdigest()[:64]


def _resolve_manual_source_name(source_url: str, supplied_name: str | None) -> str:
    if supplied_name and supplied_name.strip():
        return supplied_name.strip()
    hostname = urlparse(source_url).netloc
    return hostname or MANUAL_SOURCE_NAME


def _get_or_create_manual_source(db: Session) -> Source:
    source = db.scalar(select(Source).where(Source.name == MANUAL_SOURCE_NAME))
    if source:
        return source

    source = Source(
        name=MANUAL_SOURCE_NAME,
        root_url=MANUAL_SOURCE_URL,
        base_url=MANUAL_SOURCE_URL,
        country="Bangladesh",
        source_type="job_board",
        ingestion_mode="manual",
        connector_key="linkout_only",
        trust_level="unknown",
        compliance_status="manual_review_required",
        crawl_frequency="manual",
        enabled=False,
        requires_admin_review=True,
    )
    db.add(source)
    db.flush()
    return source


def _apply_manual_extraction(
    draft: Opportunity,
    *,
    extraction,
    requested_opportunity_type: str,
) -> None:
    if extraction.record_type == "unknown":
        return

    draft.title = extraction.title or draft.title
    draft.title_bn = extraction.title_bn or draft.title_bn
    draft.summary = extraction.summary or draft.summary
    draft.summary_bn = extraction.summary_bn or draft.summary_bn
    draft.summary_en = extraction.summary_en or draft.summary_en or extraction.summary
    draft.country = extraction.country or draft.country
    draft.employer_or_organization = extraction.employer or extraction.organization or draft.employer_or_organization
    draft.employer = extraction.employer or draft.employer
    draft.organization = extraction.organization or draft.organization
    draft.sector = extraction.sector or draft.sector
    draft.degree_level = extraction.degree_level or draft.degree_level
    draft.salary_min = extraction.salary_min
    draft.salary_max = extraction.salary_max
    draft.salary_currency = extraction.salary_currency or draft.salary_currency
    draft.deadline = parse_deadline(extraction.deadline_text) or draft.deadline
    draft.application_url = extraction.application_url or draft.application_url
    draft.eligibility_text = extraction.eligibility_text or draft.eligibility_text
    draft.visa_support = extraction.visa_support
    draft.journey_steps = extraction.journey_steps
    draft.documents_needed = extraction.documents_needed
    draft.typical_salary_bdt = extraction.typical_salary_bdt
    draft.language_requirements_json = {"items": extraction.language_requirements}
    draft.requirements_json = {"items": extraction.requirements}
    draft.benefits_json = {"items": extraction.benefits}
    draft.record_type = extraction.record_type
    draft.extracted_json = extraction.model_dump(mode="json")
    draft.extraction_confidence = extraction.extraction_confidence
    draft.opportunity_type = requested_opportunity_type or draft.opportunity_type


def _clean_manual_text(value: str | None) -> str | None:
    cleaned = (value or "").strip()
    return cleaned or None


def _clean_manual_list(values: list[str] | None) -> list[str] | None:
    if values is None:
        return None
    cleaned = [item.strip() for item in values if item and item.strip()]
    return cleaned


def _apply_manual_optional_fields(draft: Opportunity, payload: ManualJobEntryRequest) -> None:
    title_bn = _clean_manual_text(payload.title_bn)
    summary_bn = _clean_manual_text(payload.summary_bn)
    summary_en = _clean_manual_text(payload.summary_en)
    sector = _clean_manual_text(payload.sector)
    degree_level = _clean_manual_text(payload.degree_level)
    salary_currency = _clean_manual_text(payload.salary_currency)
    application_url = _clean_manual_text(payload.application_url)
    eligibility_text = _clean_manual_text(payload.eligibility_text)
    eligibility_text_bn = _clean_manual_text(payload.eligibility_text_bn)
    application_process = _clean_manual_text(payload.application_process)
    application_process_bn = _clean_manual_text(payload.application_process_bn)
    requirements = _clean_manual_list(payload.requirements)
    benefits = _clean_manual_list(payload.benefits)
    language_requirements = _clean_manual_list(payload.language_requirements)
    journey_steps = _clean_manual_list(payload.journey_steps)
    journey_steps_bn = _clean_manual_list(payload.journey_steps_bn)
    documents_needed = _clean_manual_list(payload.documents_needed)
    documents_needed_bn = _clean_manual_list(payload.documents_needed_bn)

    if title_bn is not None:
        draft.title_bn = title_bn
    if summary_bn is not None:
        draft.summary_bn = summary_bn
    if summary_en is not None:
        draft.summary_en = summary_en
    if sector is not None:
        draft.sector = sector
    if degree_level is not None:
        draft.degree_level = degree_level
    if payload.salary_min is not None:
        draft.salary_min = payload.salary_min
    if payload.salary_max is not None:
        draft.salary_max = payload.salary_max
    if salary_currency is not None:
        draft.salary_currency = salary_currency.upper()
    if application_url is not None:
        draft.application_url = application_url
    if eligibility_text is not None:
        draft.eligibility_text = eligibility_text
    if eligibility_text_bn is not None:
        draft.eligibility_text_bn = eligibility_text_bn
    if application_process is not None:
        draft.application_process = application_process
    if application_process_bn is not None:
        draft.application_process_bn = application_process_bn
    if payload.visa_support is not None:
        draft.visa_support = payload.visa_support
    if requirements is not None:
        draft.requirements_json = {"items": requirements}
    if benefits is not None:
        draft.benefits_json = {"items": benefits}
    if language_requirements is not None:
        draft.language_requirements_json = {"items": language_requirements}
    if journey_steps is not None:
        draft.journey_steps = journey_steps
    if journey_steps_bn is not None:
        draft.journey_steps_bn = journey_steps_bn
    if documents_needed is not None:
        draft.documents_needed = documents_needed
    if documents_needed_bn is not None:
        draft.documents_needed_bn = documents_needed_bn
    if payload.typical_salary_bdt is not None:
        draft.typical_salary_bdt = payload.typical_salary_bdt
    if payload.can_apply_from_bd is not None:
        draft.can_apply_from_bd = payload.can_apply_from_bd


def _apply_manual_eligibility(
    draft: Opportunity,
    *,
    requested_opportunity_type: str,
    extracted_json: dict | None,
    manual_can_apply_from_bd: bool | None = None,
) -> None:
    eligibility = tag_eligibility(
        source_connector_key="manual_entry",
        source_trust_level="unknown",
        record_type=(draft.record_type.value if hasattr(draft.record_type, "value") else draft.record_type) or "job",
        country=draft.country,
        eligibility_text=draft.eligibility_text,
        requirements_json=draft.requirements_json or {"items": []},
        extracted_json=extracted_json,
        title=draft.title,
        summary=draft.summary_en or draft.summary,
        employer=draft.employer,
    )
    draft.can_apply_from_bd = eligibility.can_apply_from_bd
    draft.requires_existing_work_permit = eligibility.requires_existing_work_permit
    draft.open_to_international_candidates = eligibility.open_to_international_candidates
    draft.open_to_authorized_workers_only = eligibility.open_to_authorized_workers_only
    draft.lmia_status = eligibility.lmia_status
    draft.eligibility_status = eligibility.eligibility_status
    draft.target_audience_tags = eligibility.target_audience_tags
    draft.risk_flags = eligibility.risk_flags
    draft.source_trust_badge = eligibility.source_trust_badge
    if manual_can_apply_from_bd is not None:
        draft.can_apply_from_bd = manual_can_apply_from_bd
        if manual_can_apply_from_bd:
            draft.eligibility_status = "eligible"
        elif draft.open_to_authorized_workers_only:
            draft.eligibility_status = "authorized_workers_only"
        else:
            draft.eligibility_status = "not_relevant"
    if not draft.opportunity_type:
        draft.opportunity_type = eligibility.opportunity_type or requested_opportunity_type


def _run_manual_extraction(db: Session, draft: Opportunity, requested_opportunity_type: str) -> None:
    extraction = extract_structured(db, {"title": draft.title, "body_text": draft.raw_text or ""})
    _apply_manual_extraction(draft, extraction=extraction, requested_opportunity_type=requested_opportunity_type)
    _apply_manual_eligibility(
        draft,
        requested_opportunity_type=requested_opportunity_type,
        extracted_json=extraction.model_dump(mode="json"),
    )


def _manual_bulk_bool(value: str | None) -> bool | None:
    cleaned = (value or "").strip().lower()
    if cleaned in {"", "null", "none"}:
        return None
    if cleaned in {"true", "1", "yes", "y"}:
        return True
    if cleaned in {"false", "0", "no", "n"}:
        return False
    raise ValueError(f"Invalid boolean value: {value}")


def _manual_bulk_number(value: str | None) -> float | None:
    cleaned = _clean_manual_text(value)
    if cleaned is None:
        return None
    return float(cleaned)


def _manual_bulk_int(value: str | None) -> int | None:
    cleaned = _clean_manual_text(value)
    if cleaned is None:
        return None
    return int(float(cleaned))


def _manual_bulk_list(value: str | None) -> list[str] | None:
    cleaned = _clean_manual_text(value)
    if cleaned is None:
        return None
    if "\n" in cleaned:
        items = [item.strip() for item in cleaned.splitlines() if item.strip()]
    else:
        items = [item.strip() for item in cleaned.split(",") if item.strip()]
    return items or None


def _load_bulk_upload_rows(file: UploadFile, content: bytes) -> list[dict[str, str]]:
    filename = (file.filename or "").lower()
    if filename.endswith(".xlsx"):
        try:
            import openpyxl
        except ImportError as exc:
            raise HTTPException(status_code=500, detail="openpyxl not installed") from exc
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        headers = [str(h).strip() if h is not None else "" for h in next(rows_iter, [])]
        return [
            {header: ("" if value is None else str(value).strip()) for header, value in zip(headers, row)}
            for row in rows_iter
        ]
    return [
        {str(key).strip(): ("" if value is None else str(value).strip()) for key, value in row.items()}
        for row in csv.DictReader(io.StringIO(content.decode("utf-8-sig")))
    ]


def _manual_entry_template_headers() -> list[str]:
    return [
        "title",
        "source_url",
        "raw_description",
        "source_name",
        "country",
        "employer",
        "deadline",
        "opportunity_type",
        "title_bn",
        "summary_bn",
        "summary_en",
        "sector",
        "degree_level",
        "salary_min",
        "salary_max",
        "salary_currency",
        "application_url",
        "eligibility_text",
        "eligibility_text_bn",
        "application_process",
        "application_process_bn",
        "visa_support",
        "can_apply_from_bd",
        "requirements",
        "benefits",
        "language_requirements",
        "journey_steps",
        "journey_steps_bn",
        "documents_needed",
        "documents_needed_bn",
        "typical_salary_bdt",
    ]


def _normalize_manual_bulk_row(row: dict[str, str], *, run_ai_extraction: bool) -> ManualJobEntryRequest:
    title = _clean_manual_text(row.get("title"))
    source_url = _clean_manual_text(row.get("source_url"))
    raw_description = _clean_manual_text(row.get("raw_description"))
    if not title or not source_url or not raw_description:
        raise ValueError("Missing required columns: title, source_url, raw_description")

    return ManualJobEntryRequest(
        title=title,
        source_url=source_url,
        raw_description=raw_description,
        source_name=_clean_manual_text(row.get("source_name")),
        country=_clean_manual_text(row.get("country")),
        employer=_clean_manual_text(row.get("employer")),
        deadline=_clean_manual_text(row.get("deadline")),
        opportunity_type=_clean_manual_text(row.get("opportunity_type")) or "overseas_job",
        run_ai_extraction=run_ai_extraction,
        title_bn=_clean_manual_text(row.get("title_bn")),
        summary_bn=_clean_manual_text(row.get("summary_bn")),
        summary_en=_clean_manual_text(row.get("summary_en")),
        sector=_clean_manual_text(row.get("sector")),
        degree_level=_clean_manual_text(row.get("degree_level")),
        salary_min=_manual_bulk_number(row.get("salary_min")),
        salary_max=_manual_bulk_number(row.get("salary_max")),
        salary_currency=_clean_manual_text(row.get("salary_currency")),
        application_url=_clean_manual_text(row.get("application_url")),
        eligibility_text=_clean_manual_text(row.get("eligibility_text")),
        eligibility_text_bn=_clean_manual_text(row.get("eligibility_text_bn")),
        application_process=_clean_manual_text(row.get("application_process")),
        application_process_bn=_clean_manual_text(row.get("application_process_bn")),
        visa_support=_manual_bulk_bool(row.get("visa_support")),
        can_apply_from_bd=_manual_bulk_bool(row.get("can_apply_from_bd")),
        requirements=_manual_bulk_list(row.get("requirements")),
        benefits=_manual_bulk_list(row.get("benefits")),
        language_requirements=_manual_bulk_list(row.get("language_requirements")),
        journey_steps=_manual_bulk_list(row.get("journey_steps")),
        journey_steps_bn=_manual_bulk_list(row.get("journey_steps_bn")),
        documents_needed=_manual_bulk_list(row.get("documents_needed")),
        documents_needed_bn=_manual_bulk_list(row.get("documents_needed_bn")),
        typical_salary_bdt=_manual_bulk_int(row.get("typical_salary_bdt")),
    )


def _save_manual_entry(
    db: Session,
    payload: ManualJobEntryRequest,
) -> tuple[Opportunity, bool]:
    manual_source = _get_or_create_manual_source(db)
    item_key = _manual_source_item_key(payload.source_url)
    source_name = _resolve_manual_source_name(payload.source_url, payload.source_name)

    draft = db.scalar(
        select(Opportunity).where(
            Opportunity.source_id == manual_source.id,
            Opportunity.source_item_key == item_key,
        )
    )
    if draft and draft.status == "published":
        raise HTTPException(status_code=409, detail="A published item already exists for this source URL")

    created = draft is None
    if draft is None:
        draft = Opportunity(
            source_id=manual_source.id,
            source_item_key=item_key,
            connector_key="manual_entry",
            source_name=source_name,
            source_page_url=payload.source_url,
            source_url=payload.source_url,
            original_apply_url=payload.source_url,
            content_type="manual",
            record_type="job",
            title=payload.title,
            summary=payload.raw_description[:400] or None,
            summary_en=payload.raw_description[:400] or None,
            country=payload.country,
            employer_or_organization=payload.employer,
            employer=payload.employer,
            deadline=parse_deadline(payload.deadline),
            opportunity_type=payload.opportunity_type,
            raw_text=payload.raw_description,
            application_url=payload.source_url,
            needs_admin_review=True,
            review_status="pending",
            status="pending",
            is_active=False,
            extraction_confidence=0.0,
            requirements_json={"items": []},
            language_requirements_json={"items": []},
            benefits_json={"items": []},
            journey_steps=[],
            documents_needed=[],
            target_audience_tags=[],
            risk_flags=[],
        )
        db.add(draft)
        db.flush()
    else:
        draft.source_name = source_name
        draft.source_page_url = payload.source_url
        draft.source_url = payload.source_url
        draft.original_apply_url = payload.source_url
        draft.title = payload.title
        draft.country = payload.country or draft.country
        draft.employer_or_organization = payload.employer or draft.employer_or_organization
        draft.employer = payload.employer or draft.employer
        draft.deadline = parse_deadline(payload.deadline) or draft.deadline
        draft.opportunity_type = payload.opportunity_type
        draft.raw_text = payload.raw_description
        draft.summary = payload.raw_description[:400] or draft.summary
        draft.summary_en = payload.raw_description[:400] or draft.summary_en

    draft.connector_key = "manual_entry"
    draft.content_type = "manual"
    draft.application_url = payload.source_url
    draft.needs_admin_review = True
    draft.review_status = "pending"
    draft.status = "pending"
    draft.is_active = False
    draft.reviewed_by = None
    draft.reviewed_at = None
    draft.slug = None
    draft.published_at = None

    if payload.run_ai_extraction:
        _run_manual_extraction(db, draft, payload.opportunity_type)
        _apply_manual_optional_fields(draft, payload)
        _apply_manual_eligibility(
            draft,
            requested_opportunity_type=payload.opportunity_type,
            extracted_json=draft.extracted_json,
            manual_can_apply_from_bd=payload.can_apply_from_bd,
        )
    else:
        draft.extracted_json = None
        draft.extraction_confidence = 0.0
        _apply_manual_optional_fields(draft, payload)
        _apply_manual_eligibility(
            draft,
            requested_opportunity_type=payload.opportunity_type,
            extracted_json=None,
            manual_can_apply_from_bd=payload.can_apply_from_bd,
        )

    _refresh_draft_search_tsv(db, draft.id)
    return draft, created


def _serialize_sources(db: Session, sources: list[Source]) -> list[AdminSourceOut]:
    if not sources:
        return []
    source_ids = [s.id for s in sources]

    draft_counts = {
        row.source_id: row
        for row in db.execute(
            select(
                Opportunity.source_id.label("source_id"),
                func.count(Opportunity.id).label("total"),
                func.sum(case((Opportunity.status == "pending", 1), else_=0)).label("pending"),
            )
            .where(Opportunity.source_id.in_(source_ids))
            .group_by(Opportunity.source_id)
        )
    }
    pub_counts = {
        row.source_id: int(row.count)
        for row in db.execute(
            select(
                Opportunity.source_id.label("source_id"),
                func.count(Opportunity.id).label("count"),
            )
            .where(Opportunity.source_id.in_(source_ids), Opportunity.status == "published")
            .group_by(Opportunity.source_id)
        )
    }
    raw_counts = {
        row.source_id: int(row.count)
        for row in db.execute(
            select(RawDocument.source_id.label("source_id"), func.count(RawDocument.id).label("count"))
            .where(RawDocument.source_id.in_(source_ids))
            .group_by(RawDocument.source_id)
        )
    }
    latest_run_ids = {
        row.source_id: row.max_id
        for row in db.execute(
            select(CrawlRun.source_id.label("source_id"), func.max(CrawlRun.id).label("max_id"))
            .where(CrawlRun.source_id.in_(source_ids))
            .group_by(CrawlRun.source_id)
        )
    }
    latest_runs: dict[int, CrawlRun] = {}
    if latest_run_ids:
        latest_runs = {
            r.id: r for r in db.scalars(select(CrawlRun).where(CrawlRun.id.in_(latest_run_ids.values()))).all()
        }
    # Legacy crawl job fallback
    latest_job_ids = {
        row.source_id: row.max_id
        for row in db.execute(
            select(CrawlJob.source_id.label("source_id"), func.max(CrawlJob.id).label("max_id"))
            .where(CrawlJob.source_id.in_(source_ids))
            .group_by(CrawlJob.source_id)
        )
    }
    latest_jobs: dict[int, CrawlJob] = {}
    if latest_job_ids:
        latest_jobs = {
            j.id: j for j in db.scalars(select(CrawlJob).where(CrawlJob.id.in_(latest_job_ids.values()))).all()
        }

    out: list[AdminSourceOut] = []
    for src in sources:
        dc = draft_counts.get(src.id)
        run = latest_runs.get(latest_run_ids.get(src.id))  # type: ignore[arg-type]
        job = latest_jobs.get(latest_job_ids.get(src.id))  # type: ignore[arg-type]
        # Prefer CrawlRun data; fall back to CrawlJob for legacy sources
        last_status = run.status.value if run else (job.status.value if job else None)
        last_started = run.started_at if run else (job.started_at if job else None)
        last_finished = run.finished_at if run else (job.finished_at if job else None)
        last_fetched = run.discovered_count if run else (job.pages_fetched if job else 0)
        last_extracted = run.draft_created_count if run else (job.records_extracted if job else 0)
        out.append(
            AdminSourceOut(
                id=src.id,
                name=src.name,
                base_url=src.base_url,
                root_url=src.root_url,
                country=src.country,
                country_scope=src.country_scope,
                source_type=src.source_type,
                ingestion_mode=src.ingestion_mode,
                connector_key=src.connector_key,
                trust_level=src.trust_level,
                compliance_status=src.compliance_status,
                crawl_frequency=src.crawl_frequency,
                first_crawl_mode=src.first_crawl_mode,
                feed_type=src.feed_type,
                auto_publish=src.auto_publish or False,
                is_official_seed_source=src.is_official_seed_source,
                is_deletable=src.is_deletable,
                settings_json=src.settings_json or {},
                last_status=src.last_status,
                discovered_item_count=src.discovered_item_count,
                imported_job_count=src.imported_job_count,
                skipped_item_count=src.skipped_item_count,
                needs_review_count=src.needs_review_count,
                target_audience=src.target_audience or [],
                search_keywords=src.search_keywords or [],
                enabled=src.enabled if src.enabled is not None else src.is_active,
                requires_admin_review=src.requires_admin_review,
                last_attempted_at=src.last_attempted_at,
                last_crawled_at=src.last_crawled_at,
                last_success_at=src.last_success_at,
                last_error=src.last_error,
                created_at=src.created_at,
                updated_at=src.updated_at,
                # Legacy
                source_class=src.source_class.value,
                trust_tier=src.trust_tier.value,
                access_method=src.access_method.value,
                crawl_frequency_minutes=src.crawl_frequency_minutes,
                is_active=src.is_active,
                parser_key=src.parser_key,
                search_queries=src.search_queries or [],
                search_results_limit=src.search_results_limit,
                child_page_limit=src.child_page_limit,
                page_ai_limit=src.page_ai_limit,
                max_jobs_per_page=src.max_jobs_per_page,
                # Stats
                draft_count=int(dc.total) if dc else 0,
                pending_review_count=int(dc.pending) if dc else 0,
                published_count=pub_counts.get(src.id, 0),
                raw_document_count=raw_counts.get(src.id, 0),
                last_crawl_status=last_status,
                last_crawl_started_at=last_started,
                last_crawl_finished_at=last_finished,
                last_pages_fetched=last_fetched or 0,
                last_records_extracted=last_extracted or 0,
                opportunity_count=int(dc.total) if dc else 0,
                active_opportunity_count=pub_counts.get(src.id, 0),
            )
        )
    return out


def _latest_running_crawl(db: Session, source_id: int) -> CrawlJob | None:
    return db.scalar(
        select(CrawlJob)
        .where(CrawlJob.source_id == source_id, CrawlJob.status == CrawlStatus.running)
        .order_by(CrawlJob.started_at.desc().nullslast(), CrawlJob.id.desc())
    )


def _close_stale_running_crawl(job: CrawlJob) -> bool:
    if not job.started_at:
        return False
    if job.started_at.astimezone(UTC) > datetime.now(UTC) - timedelta(minutes=STALE_RUNNING_CRAWL_MINUTES):
        return False
    job.status = CrawlStatus.failed
    job.finished_at = datetime.now(UTC)
    job.error_message = "Marked failed because the crawl was stuck in running state."
    return True


def _draft_to_review_out(draft: Opportunity, source_name: str | None) -> ReviewQueueOut:
    return ReviewQueueOut(
        id=draft.id,
        title=draft.title,
        title_bn=draft.title_bn,
        opportunity_type=draft.opportunity_type,
        source_id=draft.source_id,
        source_name=source_name or draft.source_name,
        source_page_url=draft.source_page_url or draft.source_url or "",
        document_url=draft.document_url,
        original_apply_url=draft.original_apply_url,
        content_type=draft.content_type,
        country=draft.country,
        destination_country=draft.destination_country,
        employer_or_organization=draft.employer_or_organization or draft.employer or draft.organization,
        deadline=str(draft.deadline) if draft.deadline else None,
        salary_text=draft.salary_text,
        eligibility_status=draft.eligibility_status,
        can_apply_from_bd=draft.can_apply_from_bd,
        requires_existing_work_permit=draft.requires_existing_work_permit,
        open_to_international_candidates=draft.open_to_international_candidates,
        lmia_status=draft.lmia_status,
        summary_bn=draft.summary_bn,
        summary_en=draft.summary_en or draft.summary,
        extraction_confidence=draft.extraction_confidence,
        needs_admin_review=draft.needs_admin_review,
        review_status=draft.review_status,
        reviewed_by=draft.reviewed_by,
        reviewed_at=draft.reviewed_at,
        target_audience_tags=draft.target_audience_tags or [],
        risk_flags=draft.risk_flags or [],
        source_trust_badge=draft.source_trust_badge,
        connector_key=draft.connector_key,
        admin_status=draft.admin_status,
        platform_category_bn=draft.platform_category_bn,
        platform_category_en=draft.platform_category_en,
        bangladesh_applicability=draft.bangladesh_applicability,
        bangladesh_applicability_reason=draft.bangladesh_applicability_reason,
        rural_user_fit_score=draft.rural_user_fit_score,
        actionability_score=draft.actionability_score,
        trust_score=draft.trust_score,
        overall_rank_score=draft.overall_rank_score,
        extraction_warnings=draft.extraction_warnings or [],
        raw_text=(draft.raw_text or draft.extracted_text or "")[:2000] or None,
        created_at=draft.created_at,
        field_confidences=(draft.extracted_json or {}).get("field_confidences") if isinstance(draft.extracted_json, dict) else None,
        record_type=draft.record_type.value if draft.record_type else None,
        source_url=draft.source_url,
    )


# ── Overview ───────────────────────────────────────────────────────────────────

@router.get("/overview", response_model=AdminOverviewOut)
def admin_overview(db: Session = Depends(get_db), _: User = Depends(get_admin_user)) -> AdminOverviewOut:
    now = datetime.now(UTC)
    pending_review = db.scalar(
        select(func.count()).select_from(Opportunity).where(
            Opportunity.needs_admin_review.is_(True),
            Opportunity.review_status.in_(["pending", None]),
        )
    ) or 0
    total_published = db.scalar(
        select(func.count()).select_from(Opportunity).where(Opportunity.status == "published")
    ) or 0
    stats = AdminOverviewStats(
        total_sources=db.scalar(select(func.count()).select_from(Source)) or 0,
        active_sources=db.scalar(select(func.count()).select_from(Source).where(Source.enabled.is_(True))) or 0,
        total_drafts=db.scalar(select(func.count()).select_from(Opportunity)) or 0,
        pending_review=pending_review,
        total_published=total_published,
        total_users=db.scalar(select(func.count()).select_from(User)) or 0,
        total_alert_rules=db.scalar(select(func.count()).select_from(AlertRule)) or 0,
        running_crawls=db.scalar(
            select(func.count()).select_from(CrawlRun).where(CrawlRun.status == CrawlRunStatus.running)
        ) or 0,
        failed_crawls_last_24h=db.scalar(
            select(func.count()).select_from(CrawlRun).where(
                CrawlRun.status.in_([CrawlRunStatus.failed, CrawlRunStatus.failed_config]),
                CrawlRun.started_at >= now - timedelta(hours=24),
            )
        ) or 0,
        queued_alert_events=db.scalar(
            select(func.count()).select_from(AlertEvent).where(AlertEvent.status.in_(["pending", "queued"]))
        ) or 0,
        total_opportunities=db.scalar(select(func.count()).select_from(Opportunity)) or 0,
        active_opportunities=total_published,
    )

    recent_job_rows = db.execute(
        select(CrawlJob, Source.name)
        .join(Source, Source.id == CrawlJob.source_id)
        .order_by(CrawlJob.id.desc())
        .limit(8)
    ).all()
    recent_crawls = [
        CrawlJobOut(
            id=j.id, source_id=j.source_id, source_name=n, status=j.status,
            started_at=j.started_at, finished_at=j.finished_at,
            error_message=j.error_message, pages_fetched=j.pages_fetched,
            records_extracted=j.records_extracted,
        )
        for j, n in recent_job_rows
    ]

    recent_run_rows = db.execute(
        select(CrawlRun, Source.name)
        .join(Source, Source.id == CrawlRun.source_id)
        .order_by(CrawlRun.id.desc())
        .limit(8)
    ).all()
    recent_runs = [
        CrawlRunOut(
            id=r.id, source_id=r.source_id, source_name=n,
            connector_key=r.connector_key, source_type=r.source_type,
            ingestion_mode=r.ingestion_mode, crawl_mode=r.crawl_mode,
            status=r.status, discovered_count=r.discovered_count,
            draft_created_count=r.draft_created_count, duplicate_count=r.duplicate_count,
            failed_count=r.failed_count, manual_review_count=r.manual_review_count,
            started_at=r.started_at, finished_at=r.finished_at,
            error_message=r.error_message, logs=r.logs,
        )
        for r, n in recent_run_rows
    ]

    sources = db.scalars(select(Source).order_by(Source.updated_at.desc()).limit(6)).all()
    return AdminOverviewOut(
        stats=stats,
        recent_crawls=recent_crawls,
        recent_runs=recent_runs,
        sources=_serialize_sources(db, sources),
    )


# ── AI Settings ────────────────────────────────────────────────────────────────

@router.get("/settings/ai", response_model=AdminAiSettingsOut)
def ai_settings(db: Session = Depends(get_db), _: User = Depends(get_admin_user)) -> AdminAiSettingsOut:
    return AdminAiSettingsOut(
        ai_provider=get_ai_provider(db),
        ai_api_key_configured=has_ai_api_key(db),
        ai_model=get_ai_model(db),
    )


@router.patch("/settings/ai", response_model=AdminAiSettingsOut)
def update_ai_settings(
    payload: AdminAiSettingsUpdate,
    db: Session = Depends(get_db),
    _: User = Depends(get_admin_user),
) -> AdminAiSettingsOut:
    provider = payload.ai_provider.strip().lower()
    set_setting(db, AI_PROVIDER, provider)
    if payload.ai_api_key:
        key_name = AI_API_KEY if provider not in {"groq", "mistral"} else ("mistral_api_key" if provider == "mistral" else "groq_api_key")
        set_setting(db, key_name, payload.ai_api_key.strip())
    if payload.ai_model.strip():
        set_setting(db, AI_MODEL, payload.ai_model.strip())
    return AdminAiSettingsOut(
        ai_provider=get_ai_provider(db),
        ai_api_key_configured=has_ai_api_key(db),
        ai_model=get_ai_model(db),
    )


# ── Sources ────────────────────────────────────────────────────────────────────

@router.get("/sources", response_model=list[AdminSourceOut])
def list_sources(db: Session = Depends(get_db), _: User = Depends(get_admin_user)) -> list[AdminSourceOut]:
    ensure_official_sources(db)
    sources = db.scalars(select(Source).order_by(Source.created_at.desc())).all()
    return _serialize_sources(db, sources)


@router.post("/sources", response_model=SourceOut)
def create_source(payload: SourceCreate, db: Session = Depends(get_db), _: User = Depends(get_admin_user)) -> SourceOut:
    data = payload.model_dump()
    data["base_url"] = str(payload.base_url)
    data["root_url"] = data["base_url"]
    source = Source(**data)
    db.add(source)
    db.commit()
    db.refresh(source)
    logger.info("admin_source_created", extra={"source_id": source.id, "name": source.name})
    return source


@router.patch("/sources/{source_id}", response_model=SourceOut)
def update_source(
    source_id: int,
    payload: SourceUpdate,
    db: Session = Depends(get_db),
    _: User = Depends(get_admin_user),
) -> SourceOut:
    source = db.scalar(select(Source).where(Source.id == source_id))
    if not source:
        raise HTTPException(status_code=404, detail="Source not found")
    data = payload.model_dump(exclude_unset=True)
    if payload.base_url is not None:
        data["base_url"] = str(payload.base_url)
        data.setdefault("root_url", data["base_url"])
    # Sync enabled → is_active
    if "enabled" in data:
        data["is_active"] = data["enabled"]
    for key, value in data.items():
        setattr(source, key, value)
    db.commit()
    db.refresh(source)
    logger.info("admin_source_updated", extra={"source_id": source.id})
    return source


@router.delete("/sources/{source_id}", response_model=MessageResponse)
def delete_source(source_id: int, db: Session = Depends(get_db), _: User = Depends(get_admin_user)) -> MessageResponse:
    source = db.scalar(select(Source).where(Source.id == source_id))
    if not source:
        raise HTTPException(status_code=404, detail="Source not found")
    if source.is_official_seed_source or not source.is_deletable:
        raise HTTPException(status_code=409, detail="Official seed sources cannot be deleted")
    db.delete(source)
    db.commit()
    logger.info("admin_source_deleted", extra={"source_id": source_id})
    return MessageResponse(message=f"Source {source_id} deleted")


@router.post("/sources/{source_id}/crawl", response_model=MessageResponse)
def trigger_crawl(
    source_id: int,
    force: bool = False,
    db: Session = Depends(get_db),
    _: User = Depends(get_admin_user),
) -> MessageResponse:
    from worker.tasks import run_source_crawl

    source = db.scalar(select(Source).where(Source.id == source_id))
    if not source:
        raise HTTPException(status_code=404, detail="Source not found")
    if not (source.enabled if source.enabled is not None else source.is_active):
        raise HTTPException(status_code=409, detail="Source is disabled. Enable it before crawling.")

    running = _latest_running_crawl(db, source_id)
    if running and not _close_stale_running_crawl(running):
        raise HTTPException(status_code=409, detail=f"Crawl #{running.id} is already running.")
    if running:
        db.commit()

    run_source_crawl.delay(source_id, force=force)
    logger.info("admin_crawl_triggered", extra={"source_id": source_id})
    return MessageResponse(message=f"Crawl queued for source {source_id}")


@router.post("/sources/{source_id}/test", response_model=SourceTestResult)
def test_source(source_id: int, db: Session = Depends(get_db), _: User = Depends(get_admin_user)) -> SourceTestResult:
    from app.ingestion.compliance_guard import ComplianceError, check_before_crawl
    from app.ingestion.cleaner import clean_page
    from app.ingestion.extractor import extract_jobs_structured
    from app.ingestion.parsers.registry import get_parser
    from app.ingestion.source_router import get_connector

    source = db.scalar(select(Source).where(Source.id == source_id))
    if not source:
        raise HTTPException(status_code=404, detail="Source not found")

    compliance_warning: str | None = None
    try:
        compliance_warning = check_before_crawl(source)
    except ComplianceError as exc:
        return SourceTestResult(
            source_id=source_id, source_name=source.name, pages_found=0,
            sample_titles=[],
            queries_used=[],
            search_results_found=0,
            child_pages_followed=0,
            pages_selected_for_ai=0,
            jobs_extracted_preview=0,
            compliance_warning=str(exc),
            error=None,
        )
    try:
        connector = get_connector(source)
        pages = connector.discover_items(source, crawl_mode="preview_only")
        sample_titles = [(p.title or p.url)[:120] for p in pages[:5]]
        diagnostics = connector.get_last_discovery_diagnostics()
        jobs_extracted_preview = 0
        if source.connector_key == "search_html_jobs":
            parser = get_parser(source.parser_key)
            for page in pages[:2]:
                parsed = parser(page)
                cleaned = clean_page(page)
                cleaned["title"] = parsed.get("title") or cleaned.get("title")
                jobs_extracted_preview += len(
                    extract_jobs_structured(db, cleaned, max_jobs=max(1, source.max_jobs_per_page or 10))
                )
        return SourceTestResult(
            source_id=source_id, source_name=source.name, pages_found=len(pages),
            sample_titles=sample_titles,
            queries_used=diagnostics.get("queries_used", []),
            search_results_found=int(diagnostics.get("search_results_found", 0) or 0),
            child_pages_followed=int(diagnostics.get("child_pages_followed", 0) or 0),
            pages_selected_for_ai=int(diagnostics.get("pages_selected_for_ai", len(pages)) or len(pages)),
            jobs_extracted_preview=jobs_extracted_preview,
            compliance_warning=compliance_warning,
            error=None,
        )
    except Exception as exc:
        return SourceTestResult(
            source_id=source_id, source_name=source.name, pages_found=0,
            sample_titles=[],
            queries_used=[],
            search_results_found=0,
            child_pages_followed=0,
            pages_selected_for_ai=0,
            jobs_extracted_preview=0,
            compliance_warning=compliance_warning,
            error=str(exc)[:500],
        )


@router.post("/sources/probe", response_model=SourceProbeResult)
async def probe_source(
    payload: SourceProbeRequest,
    _: User = Depends(get_admin_user),
) -> SourceProbeResult:
    """Probe a URL to auto-detect feed type and return sample titles."""
    import httpx
    from xml.etree import ElementTree as ET

    url = payload.url.strip()
    feed_type = "html"
    suggested_name: str | None = None
    sample_titles: list[str] = []
    detected_language: str | None = None
    scrapeability = check_scrapeability(url)

    try:
        async with httpx.AsyncClient(timeout=10, follow_redirects=True) as client:
            resp = await client.get(url, headers={"User-Agent": "SudokkhoBot/1.0"})
            resp.raise_for_status()

        content_type = resp.headers.get("content-type", "").lower()
        body = resp.text

        # Detect PDF
        if url.lower().endswith(".pdf") or "application/pdf" in content_type:
            feed_type = "pdf"
        # Detect RSS/Atom by content-type
        elif any(t in content_type for t in ("rss", "atom", "xml")):
            feed_type = "rss"
        else:
            # Try parsing as XML
            try:
                root = ET.fromstring(resp.content)
                tag = root.tag.lower()
                if "rss" in tag or "feed" in tag or "channel" in tag:
                    feed_type = "rss"
            except ET.ParseError:
                pass

            # HTML: check for RSS autodiscovery link
            if feed_type == "html" and 'type="application/rss+xml"' in body:
                feed_type = "rss"

        if feed_type == "rss":
            try:
                root = ET.fromstring(resp.content)
                ns = {"atom": "http://www.w3.org/2005/Atom"}
                # RSS 2.0
                for item in root.findall(".//item")[:3]:
                    t = item.findtext("title")
                    if t:
                        sample_titles.append(t.strip())
                # Atom
                for entry in root.findall(".//atom:entry", ns)[:3]:
                    t = entry.findtext("atom:title", namespaces=ns)
                    if t:
                        sample_titles.append(t.strip())
                # Channel title
                ch_title = root.findtext(".//channel/title") or root.findtext(".//atom:title", namespaces=ns)
                if ch_title:
                    suggested_name = ch_title.strip()[:120]
            except ET.ParseError:
                pass
        elif feed_type == "html":
            import re
            title_match = re.search(r"<title[^>]*>(.*?)</title>", body, re.IGNORECASE | re.DOTALL)
            if title_match:
                suggested_name = re.sub(r"<[^>]+>", "", title_match.group(1)).strip()[:120]
            og_match = re.search(r'<meta[^>]+property="og:title"[^>]+content="([^"]+)"', body, re.IGNORECASE)
            if og_match:
                suggested_name = og_match.group(1).strip()[:120]

        if "বাংলা" in body or "বি" in body[:500]:
            detected_language = "bn"
        elif body[:100].isascii():
            detected_language = "en"

    except Exception as exc:
        return SourceProbeResult(
            url=url,
            feed_type="html",
            is_scrapable=scrapeability.is_scrapable,
            scrape_warning=None if scrapeability.is_scrapable else scrapeability.reason,
            error=str(exc)[:300],
        )

    # --- ISC sector suggestion via keyword matching ---
    suggested_isc_sector: str | None = None
    isc_keyword_map = {
        "construction_isc":    ["construction", "civil", "mason", "welder", "carpenter", "plumber"],
        "ict_isc":             ["software", "developer", "IT", "tech", "digital", "programmer"],
        "agrofood_isc":        ["food", "agriculture", "farm", "fishery", "dairy"],
        "tourism_isc":         ["hotel", "hospitality", "restaurant", "tourism", "chef"],
        "rgt_isc":             ["garments", "textile", "sewing", "fabric", "apparel"],
        "leather_isc":         ["leather", "footwear", "tannery", "shoe"],
        "light_eng_isc":       ["engineering", "mechanic", "machinist", "fitter"],
        "pharma_isc":          ["pharmaceutical", "medicine", "laboratory", "pharmacy"],
        "furniture_isc":       ["furniture", "carpentry", "wood", "cabinet"],
        "agriculture_isc":     ["agriculture", "farming", "crop", "livestock", "poultry"],
        "informal_isc":        ["general labor", "helper", "driver", "cleaner", "domestic"],
    }
    if suggested_name or sample_titles:
        combined_probe_text = " ".join(filter(None, [suggested_name] + sample_titles)).lower()
        best_sector_hits = 0
        for sector_key, kws in isc_keyword_map.items():
            hits = sum(1 for kw in kws if kw.lower() in combined_probe_text)
            if hits > best_sector_hits:
                best_sector_hits = hits
                suggested_isc_sector = sector_key

    # --- Estimated opportunities count (links found, max 10) ---
    estimated_opportunities_per_crawl: int | None = None
    if feed_type == "html":
        import re as _re
        link_count = len(_re.findall(r'href=["\']([^"\']+)["\']', body))
        estimated_opportunities_per_crawl = min(link_count, 50)
    elif feed_type == "rss":
        estimated_opportunities_per_crawl = len(sample_titles)

    return SourceProbeResult(
        url=url,
        feed_type=feed_type,
        suggested_name=suggested_name,
        sample_titles=sample_titles[:3],
        detected_language=detected_language,
        suggested_isc_sector=suggested_isc_sector,
        estimated_opportunities_per_crawl=estimated_opportunities_per_crawl,
        is_scrapable=scrapeability.is_scrapable,
        scrape_warning=None if scrapeability.is_scrapable else scrapeability.reason,
    )


# ── Crawl jobs / runs ─────────────────────────────────────────────────────────

@router.get("/crawl-jobs", response_model=CrawlJobPage)
def crawl_jobs(
    page: int = 1, page_size: int = 20,
    db: Session = Depends(get_db), _: User = Depends(get_admin_user),
) -> CrawlJobPage:
    stmt = select(CrawlJob, Source.name).join(Source, Source.id == CrawlJob.source_id).order_by(CrawlJob.id.desc())
    total = db.scalar(select(func.count()).select_from(stmt.subquery())) or 0
    items = db.execute(stmt.offset((page - 1) * page_size).limit(page_size)).all()
    return CrawlJobPage(
        items=[
            CrawlJobOut(
                id=j.id, source_id=j.source_id, source_name=n, status=j.status,
                started_at=j.started_at, finished_at=j.finished_at,
                error_message=j.error_message, pages_fetched=j.pages_fetched,
                records_extracted=j.records_extracted,
            )
            for j, n in items
        ],
        total=total, page=page, page_size=page_size,
    )


@router.get("/crawl-runs", response_model=CrawlRunPage)
def crawl_runs(
    page: int = 1, page_size: int = 20, source_id: int | None = None,
    db: Session = Depends(get_db), _: User = Depends(get_admin_user),
) -> CrawlRunPage:
    stmt = select(CrawlRun, Source.name).join(Source, Source.id == CrawlRun.source_id).order_by(CrawlRun.id.desc())
    if source_id:
        stmt = stmt.where(CrawlRun.source_id == source_id)
    total = db.scalar(select(func.count()).select_from(stmt.subquery())) or 0
    items = db.execute(stmt.offset((page - 1) * page_size).limit(page_size)).all()
    return CrawlRunPage(
        items=[
            CrawlRunOut(
                id=r.id, source_id=r.source_id, source_name=n,
                connector_key=r.connector_key, source_type=r.source_type,
                ingestion_mode=r.ingestion_mode, crawl_mode=r.crawl_mode,
                status=r.status, discovered_count=r.discovered_count,
                parsed_count=r.parsed_count, duplicate_count=r.duplicate_count,
                draft_created_count=r.draft_created_count, draft_updated_count=r.draft_updated_count,
                unchanged_count=r.unchanged_count, skipped_count=r.skipped_count,
                failed_count=r.failed_count, manual_review_count=r.manual_review_count,
                started_at=r.started_at, finished_at=r.finished_at,
                error_message=r.error_message, logs=r.logs,
            )
            for r, n in items
        ],
        total=total, page=page, page_size=page_size,
    )


@router.post("/crawls/trigger-all", response_model=TriggerAllResult)
def trigger_all_crawls(db: Session = Depends(get_db), _: User = Depends(get_admin_user)) -> TriggerAllResult:
    from worker.tasks import run_source_crawl

    sources = db.scalars(
        select(Source).where(Source.enabled.is_(True))
    ).all()
    queued = 0
    skipped = 0
    for source in sources:
        running = _latest_running_crawl(db, source.id)
        if running and not _close_stale_running_crawl(running):
            skipped += 1
            continue
        run_source_crawl.delay(source.id)
        queued += 1
    db.commit()
    return TriggerAllResult(queued=queued, skipped=skipped)


# ── Review queue ───────────────────────────────────────────────────────────────

@router.get("/review-queue", response_model=ReviewQueuePage)
def review_queue(
    page: int = 1, page_size: int = 20,
    status: str = "pending",
    db: Session = Depends(get_db), _: User = Depends(get_admin_user),
) -> ReviewQueuePage:
    stmt = (
        select(Opportunity, Source.name)
        .join(Source, Source.id == Opportunity.source_id)
        .where(
            Opportunity.needs_admin_review.is_(True),
            Opportunity.review_status.in_([status, None] if status == "pending" else [status]),
        )
        .order_by(Opportunity.created_at.desc())
    )
    total = db.scalar(select(func.count()).select_from(stmt.subquery())) or 0
    rows = db.execute(stmt.offset((page - 1) * page_size).limit(page_size)).all()
    return ReviewQueuePage(
        items=[_draft_to_review_out(draft, src_name) for draft, src_name in rows],
        total=total, page=page, page_size=page_size,
    )


@router.post("/review/{draft_id}/approve", response_model=ReviewQueueOut)
def approve_draft(
    draft_id: int,
    db: Session = Depends(get_db),
    admin: User = Depends(get_admin_user),
) -> ReviewQueueOut:
    """
    Approve a draft — sets status='published' directly on the Opportunity row.
    This is the ONLY path that makes content visible to public users.
    """
    draft = db.scalar(select(Opportunity).where(Opportunity.id == draft_id))
    if not draft:
        raise HTTPException(status_code=404, detail="Draft not found")

    draft.status = "published"
    draft.review_status = "approved"
    draft.admin_status = "auto_approved"
    draft.is_active = True
    draft.reviewed_by = admin.id
    draft.reviewed_at = datetime.now(UTC)
    draft.published_at = draft.published_at or datetime.now(UTC)
    draft.slug = draft.slug or (_slugify(draft.title or "opportunity") + f"-{draft_id}")

    db.flush()

    _refresh_draft_search_tsv(db, draft.id)

    db.commit()
    source_name = db.scalar(select(Source.name).where(Source.id == draft.source_id))
    logger.info("admin_draft_approved", extra={"draft_id": draft_id})
    return _draft_to_review_out(draft, source_name)


@router.post("/review/{draft_id}/reject", response_model=ReviewQueueOut)
def reject_draft(
    draft_id: int,
    payload: ReviewStatusUpdate,
    db: Session = Depends(get_db),
    admin: User = Depends(get_admin_user),
) -> ReviewQueueOut:
    draft = db.scalar(select(Opportunity).where(Opportunity.id == draft_id))
    if not draft:
        raise HTTPException(status_code=404, detail="Draft not found")

    draft.status = "rejected"
    draft.review_status = "rejected"
    draft.admin_status = "rejected"
    draft.is_active = False
    draft.reviewed_by = admin.id
    draft.reviewed_at = datetime.now(UTC)
    db.commit()
    source_name = db.scalar(select(Source.name).where(Source.id == draft.source_id))
    logger.info("admin_draft_rejected", extra={"draft_id": draft_id})
    return _draft_to_review_out(draft, source_name)


@router.post("/review/{draft_id}/needs-manual-fix", response_model=ReviewQueueOut)
def needs_manual_fix(
    draft_id: int,
    payload: ReviewStatusUpdate,
    db: Session = Depends(get_db),
    admin: User = Depends(get_admin_user),
) -> ReviewQueueOut:
    draft = db.scalar(select(Opportunity).where(Opportunity.id == draft_id))
    if not draft:
        raise HTTPException(status_code=404, detail="Draft not found")

    draft.status = "pending"
    draft.review_status = "needs_manual_fix"
    draft.admin_status = "needs_review"
    draft.is_active = False
    draft.reviewed_by = admin.id
    draft.reviewed_at = datetime.now(UTC)
    db.commit()
    source_name = db.scalar(select(Source.name).where(Source.id == draft.source_id))
    logger.info("admin_draft_needs_fix", extra={"draft_id": draft_id})
    return _draft_to_review_out(draft, source_name)


@router.post("/review/{draft_id}/translate", response_model=ReviewQueueOut)
def translate_draft(
    draft_id: int,
    db: Session = Depends(get_db),
    admin: User = Depends(get_admin_user),
) -> ReviewQueueOut:
    """Fill missing title_bn / summary_bn / summary_en via LLM translation."""
    draft = db.scalar(select(Opportunity).where(Opportunity.id == draft_id))
    if not draft:
        raise HTTPException(status_code=404, detail="Draft not found")

    needs_translation = not draft.title_bn or not draft.summary_bn or not draft.summary_en
    if not needs_translation:
        source_name = db.scalar(select(Source.name).where(Source.id == draft.source_id))
        return _draft_to_review_out(draft, source_name)

    api_key = get_ai_api_key(db)
    if not api_key:
        raise HTTPException(status_code=503, detail="AI API key not configured")

    provider = get_ai_provider(db)
    source_title = draft.title or ""
    source_summary = draft.summary_en or draft.summary_bn or ""

    translate_prompt = (
        "You are a bilingual translator for a Bangladeshi job/opportunity platform.\n"
        "Given the following title and summary in English (or mixed language), produce:\n"
        "1. title_bn: Bengali translation of the title\n"
        "2. summary_bn: Bengali summary (2–4 sentences) suitable for Bangladeshi migrant workers\n"
        "3. summary_en: Clear English summary (2–4 sentences)\n\n"
        f"Title: {source_title}\n"
        f"Summary/Content: {source_summary[:3000]}\n\n"
        "Respond ONLY with valid JSON: "
        '{"title_bn": "...", "summary_bn": "...", "summary_en": "..."}'
    )

    try:
        import json as _json

        if provider == "mistral":
            from mistralai import Mistral as _Mistral
            client = _Mistral(api_key=api_key)
            resp = client.chat.complete(
                model=get_ai_model(db),
                messages=[{"role": "user", "content": translate_prompt}],
            )
            raw = resp.choices[0].message.content or "{}"
        else:
            from langchain_groq import ChatGroq as _ChatGroq
            model = _ChatGroq(model=get_ai_model(db), api_key=api_key, temperature=0.0)
            raw = model.invoke(translate_prompt).content

        translated = _json.loads(raw)
        if not draft.title_bn:
            draft.title_bn = translated.get("title_bn") or draft.title_bn
        if not draft.summary_bn:
            draft.summary_bn = translated.get("summary_bn") or draft.summary_bn
        if not draft.summary_en:
            draft.summary_en = translated.get("summary_en") or draft.summary_en

        db.commit()
        logger.info("admin_draft_translated", extra={"draft_id": draft_id})
    except Exception as exc:
        logger.warning("translate_draft_failed", extra={"draft_id": draft_id, "error": str(exc)})
        raise HTTPException(status_code=502, detail=f"Translation failed: {exc}") from exc

    source_name = db.scalar(select(Source.name).where(Source.id == draft.source_id))
    return _draft_to_review_out(draft, source_name)


@router.delete("/review/{draft_id}", response_model=MessageResponse)
def delete_review_draft(
    draft_id: int,
    db: Session = Depends(get_db),
    _: User = Depends(get_admin_user),
) -> MessageResponse:
    """Remove a draft opportunity from the review queue."""
    draft = db.scalar(select(Opportunity).where(Opportunity.id == draft_id))
    if not draft:
        raise HTTPException(status_code=404, detail="Draft not found")
    source = db.scalar(select(Source).where(Source.id == draft.source_id))
    if source and source.is_official_seed_source:
        raise HTTPException(status_code=409, detail="Crawled jobs from official sources cannot be hard deleted; archive, hide, reject, or mark inactive instead.")

    db.delete(draft)
    db.commit()
    logger.info("admin_draft_deleted", extra={"draft_id": draft_id})
    return MessageResponse(message=f"Draft {draft_id} deleted")


@router.patch("/opportunities/{opportunity_id}/review-status", response_model=ReviewQueueOut)
def set_review_status(
    opportunity_id: int,
    payload: ReviewStatusUpdate,
    db: Session = Depends(get_db),
    admin: User = Depends(get_admin_user),
) -> ReviewQueueOut:
    """Legacy single-endpoint for approve/reject/needs_manual_fix (used by old frontend)."""
    valid = {"approved", "rejected", "needs_manual_fix", "hidden", "archived", "inactive", "needs_review", "auto_approved"}
    if payload.status not in valid:
        raise HTTPException(status_code=422, detail=f"status must be one of: {valid}")

    if payload.status == "approved":
        return approve_draft(opportunity_id, db=db, admin=admin)
    elif payload.status == "rejected":
        return reject_draft(opportunity_id, payload=payload, db=db, admin=admin)
    elif payload.status == "needs_manual_fix":
        return needs_manual_fix(opportunity_id, payload=payload, db=db, admin=admin)
    draft = db.scalar(select(Opportunity).where(Opportunity.id == opportunity_id))
    if not draft:
        raise HTTPException(status_code=404, detail="Draft not found")
    draft.admin_status = "needs_review" if payload.status == "needs_review" else payload.status
    draft.needs_admin_review = payload.status == "needs_review"
    if payload.status == "auto_approved":
        draft.status = "published"
        draft.review_status = "approved"
        draft.is_active = True
        draft.published_at = draft.published_at or datetime.now(UTC)
    elif payload.status in {"hidden", "archived", "inactive"}:
        draft.is_active = False
        if draft.status == "published":
            draft.status = "expired"
    draft.reviewed_by = admin.id
    draft.reviewed_at = datetime.now(UTC)
    db.commit()
    source_name = db.scalar(select(Source.name).where(Source.id == draft.source_id))
    return _draft_to_review_out(draft, source_name)


@router.patch("/review/{draft_id}", response_model=ReviewQueueOut)
def edit_draft(
    draft_id: int,
    payload: dict,
    db: Session = Depends(get_db),
    _: User = Depends(get_admin_user),
) -> ReviewQueueOut:
    """Allow admin to correct fields before approving."""
    draft = db.scalar(select(Opportunity).where(Opportunity.id == draft_id))
    if not draft:
        raise HTTPException(status_code=404, detail="Draft not found")

    editable = {
        "title", "title_bn", "summary_bn", "summary_en",
        "country", "destination_country", "employer_or_organization",
        "job_title", "deadline", "salary_text", "eligibility_text",
        "lmia_status", "can_apply_from_bd", "eligibility_status",
        "requires_existing_work_permit", "open_to_international_candidates",
        "visa_or_work_permit_info", "application_process", "raw_text",
    }
    for field, value in payload.items():
        if field in editable:
            setattr(draft, field, value)

    db.commit()
    source_name = db.scalar(select(Source.name).where(Source.id == draft.source_id))
    return _draft_to_review_out(draft, source_name)


class ReviewDraftPatch(BaseModel):
    """Partial update of a draft from the inline editor in the review queue.

    All fields optional — the reviewer typically only changes a few. Only the
    listed fields are accepted; admin-only metadata (status, review_status,
    reviewed_by) is set via the existing approve/reject/needs-fix endpoints.
    """
    title: str | None = None
    title_bn: str | None = None
    title_en: str | None = None
    summary: str | None = None
    summary_bn: str | None = None
    summary_en: str | None = None
    job_title: str | None = None
    job_title_bn: str | None = None
    job_title_en: str | None = None
    country: str | None = None
    destination_country: str | None = None
    employer_or_organization: str | None = None
    employer: str | None = None
    organization: str | None = None
    sector: str | None = None
    deadline: str | None = None
    opportunity_type: str | None = None
    salary_min: float | None = None
    salary_max: float | None = None
    salary_currency: str | None = None
    salary_text: str | None = None
    salary_text_bn: str | None = None
    salary_text_en: str | None = None
    location_text: str | None = None
    application_url: str | None = None
    eligibility_text: str | None = None
    eligibility_text_bn: str | None = None
    eligibility_text_en: str | None = None
    visa_or_work_permit_info: str | None = None
    visa_or_work_permit_info_bn: str | None = None
    visa_or_work_permit_info_en: str | None = None
    education_requirement: str | None = None
    experience_requirement: str | None = None
    language_requirement: str | None = None
    application_process: str | None = None
    visa_support: bool | None = None
    can_apply_from_bd: bool | None = None
    journey_steps: list[str] | None = None
    journey_steps_bn: list[str] | None = None
    journey_steps_en: list[str] | None = None
    documents_needed: list[str] | None = None
    documents_needed_bn: list[str] | None = None
    documents_needed_en: list[str] | None = None
    typical_salary_bdt: int | None = None


_REVIEW_PATCH_DEADLINE_FIELD = "deadline"


@router.patch("/review/{draft_id}", response_model=ReviewQueueOut)
def patch_review_draft(
    draft_id: int,
    payload: ReviewDraftPatch,
    db: Session = Depends(get_db),
    admin: User = Depends(get_admin_user),
) -> ReviewQueueOut:
    """Inline edit of a pending draft from the review queue. Reviewer fixes
    fields in place and approves without going back to the manual-entry form."""
    draft = db.scalar(select(Opportunity).where(Opportunity.id == draft_id))
    if not draft:
        raise HTTPException(status_code=404, detail="Draft not found")

    # Apply only fields the reviewer explicitly set. Pydantic exclude_unset
    # gives us exactly those (None vs not-provided are different here).
    changes = payload.model_dump(exclude_unset=True)
    for field, value in changes.items():
        if field == _REVIEW_PATCH_DEADLINE_FIELD and isinstance(value, str):
            draft.deadline = parse_deadline(value)
        else:
            setattr(draft, field, value)

    draft.reviewed_by = admin.id
    draft.reviewed_at = datetime.now(UTC)
    db.commit()
    db.refresh(draft)
    source_name = db.scalar(select(Source.name).where(Source.id == draft.source_id))
    logger.info("admin_draft_patched", extra={"draft_id": draft_id, "fields": list(changes.keys())})
    return _draft_to_review_out(draft, source_name)


class TranslateFieldRequest(BaseModel):
    text: str
    source_lang: str = "en"  # 'bn' or 'en'
    target_lang: str = "bn"
    field_name: str | None = None  # optional, helps the LLM pick the right register


class TranslateFieldResponse(BaseModel):
    translation: str


@router.post("/translate-field", response_model=TranslateFieldResponse)
def translate_field(
    payload: TranslateFieldRequest,
    db: Session = Depends(get_db),
    _: User = Depends(get_admin_user),
) -> TranslateFieldResponse:
    """Translate a single text field (used by per-field translate buttons in
    the manual entry form and the review queue inline editor)."""
    if payload.source_lang not in ("bn", "en") or payload.target_lang not in ("bn", "en"):
        raise HTTPException(status_code=422, detail="source_lang and target_lang must be 'bn' or 'en'")
    if payload.source_lang == payload.target_lang:
        return TranslateFieldResponse(translation=payload.text)

    from app.services.translation_service import translate_text as _translate_text
    try:
        translated = _translate_text(
            db,
            text=payload.text,
            source_lang=payload.source_lang,
            target_lang=payload.target_lang,
            field_hint=payload.field_name,
        )
    except RuntimeError as exc:
        raise HTTPException(status_code=503, detail=str(exc)) from exc
    except Exception as exc:
        logger.warning("translate_field_failed", extra={"error": str(exc)})
        raise HTTPException(status_code=502, detail=f"Translation failed: {exc}") from exc
    return TranslateFieldResponse(translation=translated)


class DuplicateCheckResponse(BaseModel):
    found: bool
    draft_id: int | None = None
    status: str | None = None  # 'pending' | 'published' | 'rejected' | 'expired'
    title: str | None = None
    created_at: datetime | None = None


@router.get("/check-duplicate", response_model=DuplicateCheckResponse)
def check_duplicate_url(
    url: str,
    db: Session = Depends(get_db),
    _: User = Depends(get_admin_user),
) -> DuplicateCheckResponse:
    """Look up an existing draft or published opportunity by source URL.
    Used by the manual entry form to warn the reviewer before they submit a
    duplicate."""
    url_clean = (url or "").strip()
    if not url_clean:
        return DuplicateCheckResponse(found=False)
    existing = db.scalar(
        select(Opportunity)
        .where(
            (Opportunity.source_page_url == url_clean)
            | (Opportunity.source_url == url_clean)
            | (Opportunity.application_url == url_clean)
        )
        .order_by(Opportunity.id.desc())
        .limit(1)
    )
    if not existing:
        return DuplicateCheckResponse(found=False)
    return DuplicateCheckResponse(
        found=True,
        draft_id=existing.id,
        status=existing.status,
        title=existing.title,
        created_at=existing.created_at,
    )


class DiscoveryRunRequest(BaseModel):
    query: str = Field(min_length=2, max_length=500)
    target_country: str | None = None
    max_results: int = Field(default=12, ge=1, le=30)


class DiscoveryRunDraft(BaseModel):
    draft_id: int
    title: str
    url: str
    confidence: float
    is_new: bool


class DiscoveryRunResponse(BaseModel):
    query: str
    variants: list[str]
    urls_considered: int
    drafts_created: int
    drafts_updated: int
    duplicates: int
    failed: int
    drafts: list[DiscoveryRunDraft]
    warnings: list[str]


@router.post("/discover", response_model=DiscoveryRunResponse)
def run_discovery_endpoint(
    payload: DiscoveryRunRequest,
    db: Session = Depends(get_db),
    _: User = Depends(get_admin_user),
) -> DiscoveryRunResponse:
    """Agentic discovery: plan → search → score → ingest. Drafts land in the
    review queue under the synthetic 'Agentic Discovery' source."""
    from app.ingestion.discovery_agent import run_discovery as _run

    report = _run(
        db,
        query=payload.query,
        target_country=payload.target_country,
        max_results=payload.max_results,
    )
    return DiscoveryRunResponse(
        query=report.query,
        variants=report.variants,
        urls_considered=report.urls_considered,
        drafts_created=report.drafts_created,
        drafts_updated=report.drafts_updated,
        duplicates=report.duplicates,
        failed=report.failed,
        drafts=[DiscoveryRunDraft(
            draft_id=d.draft_id,
            title=d.title,
            url=d.url,
            confidence=d.confidence,
            is_new=d.is_new,
        ) for d in report.drafts],
        warnings=report.warnings,
    )


@router.post("/manual-entry/extract-from-file", response_model=ManualEntryPrefillResponse)
def extract_from_file(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _: User = Depends(get_admin_user),
) -> ManualEntryPrefillResponse:
    """Upload a job-poster image or PDF, run OCR + extraction, return form-ready
    data. Returns the same shape as `/manual-entry/prefill` so the frontend can
    reuse the same merge logic."""
    from app.ingestion.cleaner import clean_page as _clean_page
    from app.ingestion.pdf_extractor import extract_text as _pdf_extract
    from app.ingestion.schemas import FetchedPage as _FetchedPage

    raw = file.file.read()
    if not raw:
        raise HTTPException(status_code=400, detail="Empty file")
    if len(raw) > 20 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="File exceeds 20 MB limit")

    filename = file.filename or "upload"
    ctype = (file.content_type or "").lower()
    warnings: list[str] = []
    extracted_text = ""

    if "pdf" in ctype or filename.lower().endswith(".pdf"):
        try:
            result = _pdf_extract(raw)
            extracted_text = result.text
            if result.used_ocr:
                warnings.append("PDF was OCR-scanned (no text layer); double-check extracted text.")
        except Exception as exc:
            raise HTTPException(status_code=422, detail=f"PDF parse failed: {exc}") from exc
    elif ctype.startswith("image/") or filename.lower().endswith((".jpg", ".jpeg", ".png", ".webp", ".bmp", ".tiff")):
        try:
            import io as _io
            from PIL import Image as _Image
            import pytesseract as _ocr
            image = _Image.open(_io.BytesIO(raw))
            extracted_text = _ocr.image_to_string(image, lang="ben+eng").strip()
            warnings.append("Image OCR used — verify all extracted fields.")
        except Exception as exc:
            raise HTTPException(status_code=422, detail=f"Image OCR failed: {exc}") from exc
    else:
        raise HTTPException(status_code=415, detail="Unsupported file type — use PDF or image")

    if not extracted_text.strip():
        raise HTTPException(status_code=422, detail="No text extracted from file")

    # Run the structured extractor over the OCR/PDF text.
    page = _FetchedPage(
        url=f"upload://{filename}",
        title=filename[:200],
        raw_text=extracted_text,
        content_type="manual",
    )
    cleaned = _clean_page(page)
    extraction = extract_structured(db, cleaned)
    if extraction.record_type == "unknown":
        warnings.append("Extractor could not classify this file as an opportunity. Review carefully.")

    data = extraction.model_dump()
    return ManualEntryPrefillResponse(
        title=data.get("title") or filename,
        title_bn=data.get("title_bn"),
        summary_en=data.get("summary_en") or data.get("summary"),
        summary_bn=data.get("summary_bn"),
        raw_description=extracted_text[:6000].strip() or None,
        country=data.get("country"),
        employer=data.get("employer") or data.get("organization"),
        deadline=data.get("deadline_text"),
        opportunity_type=_infer_opportunity_type(extraction.record_type),
        sector=data.get("sector"),
        degree_level=data.get("degree_level"),
        salary_min=data.get("salary_min"),
        salary_max=data.get("salary_max"),
        salary_currency=data.get("salary_currency"),
        application_url=data.get("application_url"),
        eligibility_text=data.get("eligibility_text"),
        visa_support=data.get("visa_support"),
        can_apply_from_bd=data.get("can_apply_from_bd"),
        requirements=[str(r) for r in (data.get("requirements") or []) if r],
        benefits=[str(b) for b in (data.get("benefits") or []) if b],
        language_requirements=[str(l) for l in (data.get("language_requirements") or []) if l],
        journey_steps=[str(s) for s in (data.get("journey_steps") or []) if s],
        documents_needed=[str(d) for d in (data.get("documents_needed") or []) if d],
        typical_salary_bdt=data.get("typical_salary_bdt"),
        extraction_confidence=float(data.get("extraction_confidence") or 0.0),
        fetched_url=f"upload://{filename}",
        warnings=warnings,
    )


@router.post("/manual-entry/prefill", response_model=ManualEntryPrefillResponse)
def prefill_manual_entry(
    payload: ManualEntryPrefillRequest,
    db: Session = Depends(get_db),
    _: User = Depends(get_admin_user),
) -> ManualEntryPrefillResponse:
    """Fetch a URL, run the AI extractor, return form-shaped data.

    Reused infrastructure: httpx + clean_page + extract_structured. Does NOT
    write to the DB — the reviewer must still submit the form. Designed for the
    'paste URL → auto-fill' flow on the manual entry page.
    """
    import httpx as _httpx

    from app.ingestion.cleaner import clean_page as _clean_page
    from app.ingestion.schemas import FetchedPage as _FetchedPage

    url = payload.url.strip()
    warnings: list[str] = []
    raw_html = ""
    page_title: str | None = None
    try:
        resp = _httpx.get(
            url,
            timeout=30,
            follow_redirects=True,
            headers={"User-Agent": "Mozilla/5.0 (compatible; JobFinder-Prefill/1.0)"},
        )
        resp.raise_for_status()
        raw_html = resp.text
        # Best-effort title extraction (full HTML parsing happens in cleaner).
        title_match = re.search(r"<title[^>]*>(.*?)</title>", raw_html, re.IGNORECASE | re.DOTALL)
        if title_match:
            page_title = title_match.group(1).strip()[:500] or None
    except _httpx.HTTPStatusError as exc:
        raise HTTPException(status_code=502, detail=f"Source returned HTTP {exc.response.status_code}") from exc
    except _httpx.HTTPError as exc:
        raise HTTPException(status_code=502, detail=f"Failed to fetch URL: {exc}") from exc

    page = _FetchedPage(url=url, raw_html=raw_html, title=page_title, content_type="html")
    cleaned = _clean_page(page)

    if not (cleaned.get("body_text") or "").strip():
        warnings.append("Page contained no extractable body text — fill fields manually.")

    extraction = extract_structured(db, cleaned)
    if extraction.record_type == "unknown":
        warnings.append("Extractor could not classify this page as an opportunity. Review carefully.")

    data = extraction.model_dump()
    requirements = data.get("requirements") or []
    benefits = data.get("benefits") or []
    language_requirements = data.get("language_requirements") or []
    journey_steps = data.get("journey_steps") or []
    documents_needed = data.get("documents_needed") or []
    body_excerpt = (cleaned.get("body_text") or "")[:6000].strip() or None

    return ManualEntryPrefillResponse(
        title=data.get("title") or page_title,
        title_bn=data.get("title_bn"),
        summary_en=data.get("summary_en") or data.get("summary"),
        summary_bn=data.get("summary_bn"),
        raw_description=body_excerpt,
        country=data.get("country"),
        employer=data.get("employer") or data.get("organization"),
        deadline=data.get("deadline_text"),
        opportunity_type=_infer_opportunity_type(extraction.record_type),
        sector=data.get("sector"),
        degree_level=data.get("degree_level"),
        salary_min=data.get("salary_min"),
        salary_max=data.get("salary_max"),
        salary_currency=data.get("salary_currency"),
        application_url=data.get("application_url") or url,
        eligibility_text=data.get("eligibility_text"),
        visa_support=data.get("visa_support"),
        can_apply_from_bd=data.get("can_apply_from_bd"),
        requirements=[str(r) for r in requirements if r],
        benefits=[str(b) for b in benefits if b],
        language_requirements=[str(l) for l in language_requirements if l],
        journey_steps=[str(s) for s in journey_steps if s],
        documents_needed=[str(d) for d in documents_needed if d],
        typical_salary_bdt=data.get("typical_salary_bdt"),
        extraction_confidence=float(data.get("extraction_confidence") or 0.0),
        fetched_url=url,
        warnings=warnings,
    )


def _infer_opportunity_type(record_type: str | None) -> str:
    """Map extractor's record_type to the form's opportunity_type select values."""
    if record_type == "scholarship":
        return "scholarship"
    if record_type == "policy_update":
        return "overseas_job"  # policy updates aren't a form option; default sensibly
    return "overseas_job"


@router.post("/manual-entry", response_model=ReviewQueueOut)
def create_manual_entry(
    payload: ManualJobEntryRequest,
    db: Session = Depends(get_db),
    _: User = Depends(get_admin_user),
) -> ReviewQueueOut:
    draft, _created = _save_manual_entry(db, payload)
    db.commit()
    db.refresh(draft)
    logger.info("admin_manual_entry_created", extra={"draft_id": draft.id, "source_url": payload.source_url})
    return _draft_to_review_out(draft, draft.source_name)


@router.post("/manual-entry/bulk", response_model=ManualEntryBulkImportResult)
def bulk_manual_entry_import(
    file: UploadFile = File(...),
    run_ai_extraction: bool = Form(True),
    db: Session = Depends(get_db),
    _: User = Depends(get_admin_user),
) -> ManualEntryBulkImportResult:
    content = file.file.read()
    filename = (file.filename or "").lower()
    if not filename.endswith(".csv") and not filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only CSV and XLSX files are supported")

    raw_rows = _load_bulk_upload_rows(file, content)
    created = 0
    updated = 0
    skipped = 0
    draft_ids: list[int] = []
    errors: list[BulkImportError] = []

    for idx, row in enumerate(raw_rows, start=2):
        try:
            payload = _normalize_manual_bulk_row(row, run_ai_extraction=run_ai_extraction)
            draft, was_created = _save_manual_entry(db, payload)
            db.commit()
            db.refresh(draft)
            draft_ids.append(draft.id)
            if was_created:
                created += 1
            else:
                updated += 1
        except HTTPException as exc:
            db.rollback()
            skipped += 1
            errors.append(BulkImportError(row=idx, detail=str(exc.detail)))
        except Exception as exc:
            db.rollback()
            skipped += 1
            errors.append(BulkImportError(row=idx, detail=str(exc)))

    return ManualEntryBulkImportResult(
        created=created,
        updated=updated,
        skipped=skipped,
        errors=errors,
        draft_ids=draft_ids,
    )


@router.get("/manual-entry/bulk-template")
def manual_entry_bulk_template(_: User = Depends(get_admin_user)) -> Response:
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(_manual_entry_template_headers())
    return Response(
        content=output.getvalue(),
        media_type="text/csv",
        headers={"Content-Disposition": 'attachment; filename="manual-entry-template.csv"'},
    )


@router.post("/manual-entry/{draft_id}/re-extract", response_model=ReviewQueueOut)
def re_extract_manual_entry(
    draft_id: int,
    db: Session = Depends(get_db),
    _: User = Depends(get_admin_user),
) -> ReviewQueueOut:
    draft = db.scalar(select(Opportunity).where(Opportunity.id == draft_id))
    if not draft:
        raise HTTPException(status_code=404, detail="Draft not found")
    if draft.connector_key != "manual_entry":
        raise HTTPException(status_code=400, detail="Only manual-entry drafts can be re-extracted")
    if not draft.raw_text:
        raise HTTPException(status_code=400, detail="Draft has no raw text to re-extract")

    draft.needs_admin_review = True
    draft.review_status = "pending"
    draft.status = "pending"
    draft.is_active = False
    draft.reviewed_by = None
    draft.reviewed_at = None
    draft.slug = None
    draft.published_at = None

    _run_manual_extraction(db, draft, draft.opportunity_type or "overseas_job")
    _refresh_draft_search_tsv(db, draft.id)
    db.commit()
    db.refresh(draft)
    logger.info("admin_manual_entry_reextracted", extra={"draft_id": draft.id})
    return _draft_to_review_out(draft, draft.source_name)


# ── Published opportunities ────────────────────────────────────────────────────

@router.get("/published", response_model=list[ReviewQueueOut])
def list_published(
    page: int = 1, page_size: int = 20,
    db: Session = Depends(get_db), _: User = Depends(get_admin_user),
) -> list[ReviewQueueOut]:
    drafts = db.scalars(
        select(Opportunity)
        .where(Opportunity.status == "published")
        .order_by(Opportunity.published_at.desc())
        .offset((page - 1) * page_size)
        .limit(page_size)
    ).all()
    return [_draft_to_review_out(d, d.source_name) for d in drafts]


# ── Bulk import ────────────────────────────────────────────────────────────────

@router.post("/sources/bulk-import", response_model=BulkImportResult)
def bulk_import_sources(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _: User = Depends(get_admin_user),
) -> BulkImportResult:
    content = file.file.read()
    filename = (file.filename or "").lower()

    errors: list[BulkImportError] = []
    created = 0
    skipped = 0

    if filename.endswith(".xlsx"):
        try:
            import openpyxl
        except ImportError:
            raise HTTPException(status_code=500, detail="openpyxl not installed")
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        headers = [str(h).strip() if h is not None else "" for h in next(rows_iter)]
        raw_rows = [dict(zip(headers, [str(v).strip() if v is not None else "" for v in row])) for row in rows_iter]
    else:
        raw_rows = list(csv.DictReader(io.StringIO(content.decode("utf-8-sig"))))

    for idx, row in enumerate(raw_rows, start=2):
        name = row.get("name", "").strip()
        base_url = row.get("base_url", "").strip() or row.get("root_url", "").strip()
        if not name or not base_url:
            errors.append(BulkImportError(row=idx, detail="Missing required columns: name, base_url"))
            continue
        try:
            ta_raw = row.get("target_audience") or ""
            kw_raw = row.get("search_keywords") or row.get("search_queries") or ""
            payload = SourceCreate(
                name=name,
                base_url=base_url,  # type: ignore[arg-type]
                country=row.get("country") or "Bangladesh",
                source_type=row.get("source_type") or None,
                ingestion_mode=row.get("ingestion_mode") or None,
                connector_key=row.get("connector_key") or None,
                country_scope=row.get("country_scope") or None,
                target_audience=[t.strip() for t in ta_raw.split(",") if t.strip()],
                search_keywords=[k.strip() for k in kw_raw.split(",") if k.strip()],
                trust_level=row.get("trust_level") or None,
                compliance_status=row.get("compliance_status") or "unknown",
                crawl_frequency=row.get("crawl_frequency") or "daily",
                first_crawl_mode=row.get("first_crawl_mode") or "active_only",
                requires_admin_review=str(row.get("requires_admin_review", "true")).lower() in ("true", "1", "yes"),
                enabled=str(row.get("enabled", row.get("is_active", "true"))).lower() not in ("false", "0", "no"),
                search_results_limit=int(row.get("search_results_limit") or 10),
                child_page_limit=int(row.get("child_page_limit") or 10),
                page_ai_limit=int(row.get("page_ai_limit") or 25),
                max_jobs_per_page=int(row.get("max_jobs_per_page") or 10),
            )
        except Exception as exc:
            errors.append(BulkImportError(row=idx, detail=str(exc)))
            continue
        data = payload.model_dump()
        data["base_url"] = str(payload.base_url)
        data["root_url"] = data["base_url"]
        src = Source(**data)
        db.add(src)
        try:
            db.flush()
            created += 1
        except IntegrityError:
            db.rollback()
            skipped += 1

    db.commit()
    return BulkImportResult(created=created, skipped=skipped, errors=errors)


# ── Reindex / reset ────────────────────────────────────────────────────────────

@router.post("/reindex/{opportunity_id}", response_model=MessageResponse)
def reindex(opportunity_id: int, _: User = Depends(get_admin_user)) -> MessageResponse:
    from worker.tasks import reindex_opportunity
    reindex_opportunity.delay(opportunity_id)
    return MessageResponse(message=f"Reindex queued for opportunity {opportunity_id}")


@router.post("/reindex-all")
def reindex_all(db: Session = Depends(get_db), _: User = Depends(get_admin_user)) -> dict:
    from worker.tasks import reindex_opportunity
    ids = db.scalars(select(Opportunity.id)).all()
    for oid in ids:
        reindex_opportunity.delay(oid)
    return {"queued": len(ids)}


@router.post("/translate-batch")
def translate_batch(
    db: Session = Depends(get_db),
    _: User = Depends(get_admin_user),
    only_missing: bool = True,
    limit: int | None = None,
) -> dict:
    """Backfill missing _bn/_en translations on existing opportunities.

    only_missing=True (default): skip rows that already have title_bn AND title_en
    AND summary_bn AND summary_en. Pass only_missing=False to re-translate all.
    """
    from worker.tasks import translate_draft_async

    query = select(Opportunity.id)
    if only_missing:
        query = query.where(
            (Opportunity.title_bn.is_(None))
            | (Opportunity.title_en.is_(None))
            | (Opportunity.summary_bn.is_(None))
            | (Opportunity.summary_en.is_(None))
        )
    if limit:
        query = query.limit(limit)
    ids = db.scalars(query).all()
    for oid in ids:
        translate_draft_async.delay(oid, overwrite=not only_missing)
    return {"queued": len(ids), "only_missing": only_missing}


@router.post("/reset-all-data", response_model=DataResetResult)
def reset_all_data(db: Session = Depends(get_db), _: User = Depends(get_admin_user)) -> DataResetResult:
    from app.models.entities import OpportunityEmbedding

    n_emb = db.scalar(select(func.count()).select_from(OpportunityEmbedding)) or 0
    db.execute(delete(OpportunityEmbedding))
    n_pub = db.scalar(
        select(func.count()).select_from(Opportunity).where(Opportunity.status == "published")
    ) or 0
    n_opp = db.scalar(select(func.count()).select_from(Opportunity)) or 0
    db.execute(delete(Opportunity))
    n_raw = db.scalar(select(func.count()).select_from(RawDocument)) or 0
    db.execute(delete(RawDocument))
    n_runs = db.scalar(select(func.count()).select_from(CrawlRun)) or 0
    db.execute(delete(CrawlRun))
    n_jobs = db.scalar(select(func.count()).select_from(CrawlJob)) or 0
    db.execute(delete(CrawlJob))
    n_src = db.scalar(select(func.count()).select_from(Source)) or 0
    db.execute(delete(Source))
    db.commit()
    logger.warning(
        "admin_reset_all_data",
        extra={"sources": n_src, "drafts": n_opp, "published": n_pub},
    )
    return DataResetResult(
        deleted_published=n_pub, deleted_drafts=n_opp,
        deleted_raw_documents=n_raw, deleted_crawl_runs=n_runs,
        deleted_crawl_jobs=n_jobs, deleted_sources=n_src,
        deleted_opportunities=n_opp,
    )


# ── Raw documents ──────────────────────────────────────────────────────────────

@router.get("/raw-documents", response_model=RawDocumentPage)
def raw_documents(
    page: int = 1,
    page_size: int = 20,
    source_id: int | None = None,
    crawl_run_id: int | None = None,
    db: Session = Depends(get_db),
    _: User = Depends(get_admin_user),
) -> RawDocumentPage:
    stmt = select(RawDocument).order_by(RawDocument.fetched_at.desc(), RawDocument.id.desc())
    if source_id:
        stmt = stmt.where(RawDocument.source_id == source_id)
    if crawl_run_id:
        stmt = stmt.where(RawDocument.crawl_run_id == crawl_run_id)
    total = db.scalar(select(func.count()).select_from(stmt.subquery())) or 0
    items = db.scalars(stmt.offset((page - 1) * page_size).limit(page_size)).all()
    return RawDocumentPage(items=items, total=total, page=page, page_size=page_size)


@router.get("/raw-documents/{doc_id}", response_model=RawDocumentOut)
def raw_document(doc_id: int, db: Session = Depends(get_db), _: User = Depends(get_admin_user)) -> RawDocumentOut:
    doc = db.scalar(select(RawDocument).where(RawDocument.id == doc_id))
    if not doc:
        raise HTTPException(status_code=404, detail="Raw document not found")
    return doc


# ── Legacy failed-extractions endpoint (backward compat for old frontend) ──────

@router.get("/failed-extractions", response_model=FailedExtractionPage)
def failed_extractions(
    page: int = 1, page_size: int = 20,
    db: Session = Depends(get_db), _: User = Depends(get_admin_user),
) -> FailedExtractionPage:
    stmt = (
        select(Opportunity, Source.name)
        .join(Source, Source.id == Opportunity.source_id)
        .where(Opportunity.extraction_confidence < 0.45)
        .order_by(Opportunity.updated_at.desc())
    )
    total = db.scalar(select(func.count()).select_from(stmt.subquery())) or 0
    items = db.execute(stmt.offset((page - 1) * page_size).limit(page_size)).all()
    return FailedExtractionPage(
        items=[
            FailedExtractionOut(
                id=o.id, title=o.title, record_type=o.record_type,
                source_id=o.source_id, source_name=n,
                source_url=o.source_url or o.source_page_url or "",
                extraction_confidence=o.extraction_confidence, updated_at=o.updated_at,
            )
            for o, n in items
        ],
        total=total, page=page, page_size=page_size,
    )


# ── Test email endpoint ────────────────────────────────────────────────────────

class TestEmailRequest(BaseModel):
    to_email: str


class TestEmailResult(BaseModel):
    success: bool
    to_email: str
    opportunities_sent: int


@router.post("/test-email", response_model=TestEmailResult)
def test_email(
    body: TestEmailRequest,
    db: Session = Depends(get_db),
    _: User = Depends(get_admin_user),
) -> TestEmailResult:
    """Send a sample alert email with the top 3 published opportunities."""
    from app.services.email_service import send_alert_email
    from app.core.config import get_settings

    opps = db.scalars(
        select(Opportunity)
        .where(Opportunity.status == "published")
        .order_by(Opportunity.overall_rank_score.desc())
        .limit(3)
    ).all()

    settings = get_settings()
    base_url = settings.web_base_url.rstrip("/")
    opp_dicts = [
        {
            "title": opp.title,
            "title_bn": opp.title_bn,
            "country": opp.country or opp.destination_country or "",
            "deadline": str(opp.deadline) if opp.deadline else "",
            "url": f"{base_url}/opportunity/{opp.id}",
            "source_name": opp.source_name or "",
        }
        for opp in opps
    ]

    success = send_alert_email(
        to_email=body.to_email,
        user_name="Admin",
        opportunities=opp_dicts,
        locale="bn",
    )
    return TestEmailResult(success=success, to_email=body.to_email, opportunities_sent=len(opp_dicts))
